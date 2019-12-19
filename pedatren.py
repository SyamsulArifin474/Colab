import requests
import json
import aiohttp, asyncio
import base64  # , common
from requests.auth import HTTPBasicAuth
from config import conf


class Login():

    __url = 'https://api.pedatren.nuruljadid.app/'

    def __init__(self):
        # baca token.txt, jika gak ada maka buat file token.txt
        try:
            open('token.txt', 'r')
        except IOError:
            open("token.txt", "w")
        with open("token.txt", "r") as f:
            token = f.read()
        self.__token = token
        self.__url = Login.__url

    @property
    def headers(self):
        header = {
            'content-type': 'application/json',
            'connection': 'keep-alive',
            'User-Agent': conf['agent'],
            'x-token': self.__token,
        }
        return header

    @property
    def url(self):
        return self.__url

    def login(self):
        data = requests.get(self.url+'auth/login',
                            auth=(conf['user'], conf['pass']), headers=self.headers)
        if data.status_code == 200:
            self.__token = data.headers['x-token']
            with open('token.txt', 'w') as f:
                f.write(self.__token)

    def cekLogin(self):
        data = requests.get(self.url+'auth/login', headers=self.headers)
        return data.status_code

    @property
    def token(self):
        if self.cekLogin() != 200:
            self.login()
        return self.__token

    def level(self):
        user = self.token.split(".")[0]
        user += "=" * ((4 - len(user) % 4) % 4)
        level = json.loads(base64.b64decode(user))['scope'][1]
        return level

    @property
    def urlUser(self):
        lev = self.level()
        if 'lembaga' in lev:
            urlUser = "{}{}/".format(self.url, lev.replace('-', '/'))
        elif 'biktren' in lev:
            urlUser = "{}{}/".format(self.url, lev)
        else:
            urlUser = "{}/".format(self.url)
        return urlUser

    @property
    def logOut(self):
        log = requests.get(self.url+'auth/logout', headers=self.headers)
        if log.status_code == 200:
            return log.status_code


class Pedatren(Login):

    def getSantri(self, cari, lembaga=None):
        params = {
            'page': '1',
            'limit': '5',
            'wilayah': conf['wilayah'],
            'cari': cari,
            'lembaga': lembaga,
        }
        f = requests.get(self.urlUser+'santri',
                         headers=self.headers, params=params)
        f.close()
        return json.loads(f.content)

    def santri(self, cari):
        data = self.getSantri(cari)
        temp = []
        for i in data:
            tampung = {
                "uuid": i.get("uuid"),
                "nik": i.get("nik") or i.get("no_passport"),
                "nama": i.get("nama_lengkap"),
                "kabupaten": i.get("kabupaten"),
                "nis": i.get("santri").get("nis"),
            }
            if i.get("domisili_santri") != None:
                tampung["wilayah"] = i.get("domisili_santri").get("wilayah")
                tampung["blok"] = i.get("domisili_santri").get("blok")
                tampung["kamar"] = i.get("domisili_santri").get("kamar")
            if i.get("pendidikan") != None:
                tampung["lembaga"] = i.get("pendidikan").get("lembaga")
                tampung["kelas"] = i.get("pendidikan").get("kelas")
                tampung["jurusan"] = i.get("pendidikan").get("jurusan")
                tampung["rombel"] = i.get("pendidikan").get("rombel")
            temp.append(tampung)

        return temp

    def person(self, uuid):
        with requests.get(self.url+'person/{}'.format(uuid), headers=self.headers) as f:
            return json.loads(f.content)

    def getAll_santri(self, lembaga=None, daerah=None):
        # try:
        params = {
            'page': '1',
            'limit': '1000',
            'wilayah': conf['wilayah'],
            'blok': daerah,
            'lembaga': lembaga,
        }
        f = requests.get(
            self.urlUser+'santri', headers=self.headers, params=params)
        a = f.headers["x-pagination-total-page"]
        total = int(a)
        json_file = []
        for r in range(1, int(total)+1):
            params['page'] = str(r)
            data = requests.get(self.urlUser+'santri',
                                headers=self.headers, params=params)
            to_json = json.loads(data.content)
            json_file.append(to_json)
        json_file = [j for i in json_file for j in i]
        return json_file

    def all_santri(self, lembaga=None, daerah=None):
        json_file = []
        data = self.getAll_santri(lembaga=lembaga, daerah=daerah)
        for i in data:
            tampung = {
                "uuid": i.get("uuid"),
                "nik": i.get("nik") or i.get("no_passport"),
                "nama": i.get("nama_lengkap"),
                "kabupaten": i.get("kabupaten"),
                "nis": i.get("santri").get("nis"),
            }
            if i.get("domisili_santri") != None:
                tampung["wilayah"] = i.get("domisili_santri").get("wilayah")
                tampung["blok"] = i.get("domisili_santri").get("blok")
                tampung["kamar"] = i.get("domisili_santri").get("kamar")
            if i.get("pendidikan") != None:
                tampung["lembaga"] = i.get("pendidikan").get("lembaga")
                tampung["kelas"] = i.get("pendidikan").get("kelas")
                tampung["jurusan"] = i.get("pendidikan").get("jurusan")
                tampung["rombel"] = i.get("pendidikan").get("rombel")
            json_file.append(tampung)

        return json_file

    def all_pelajar(self):
        # try:
        params = {
            'page': '1',
            'limit': '1000',
            'wilayah': conf['wilayah'],
        }
        with requests.get(self.urlUser+'pelajar', headers=self.headers, params=params) as f:
            total = f.headers['x-pagination-total-page']
        json_file = []
        for r in range(1, int(total)+1):
            params['page'] = str(r)
            with requests.get(self.urlUser+'pelajar', headers=self.headers, params=params) as f:
                to_json = json.loads(f.content)
                json_file.append(to_json)
        json_file = [j for i in json_file for j in i]
        return json_file

    def catatan(self, uuid):
        afektif = requests.get(
            self.url+"person/{}/catatan/afektif".format(uuid), headers=self.headers).json()
        kognitif = requests.get(
            self.url+"person/{}/catatan/kognitif".format(uuid), headers=self.headers).json()
        afektif.extend(kognitif)
        return afektif

    def total_santri(self):
        data = requests.get(
            self.urlUser+"santri/total?wilayah=daltim", headers=self.headers).json()
        return data.get("total")

    def filter(self, pilih=None):
        data = requests.get(self.urlUser+"santri/filter",
                            headers=self.headers).json()
        # wilayah = data["domisili_santri"]
        wilayah = data["domisili_santri"][16]["list_blok"]
        lembaga = data["lembaga"]
        if pilih == "lembaga":
            return lembaga
        else:
            return wilayah

    async def getSantriPerson(self, url, session):
        async with session.get(url) as respon:
            return await respon.json()

    async def dataPelajar(self, uuid):
        tasks = []
        async with aiohttp.ClientSession(headers=self.headers) as session:
            for i in uuid:
                task = asyncio.ensure_future(self.getSantriPerson(self.url+'person/{}'.format(i), session))
                tasks.append(task)
            responses = await asyncio.gather(*tasks)
            return responses

    def simpanAllPerson(self):
        data = self.getAll_santri()
        uuid = [i['uuid'] for i in data]
        loop = asyncio.get_event_loop()
        person = loop.run_until_complete(self.dataPelajar(uuid))
        with open("dataJSON.json", "w", encoding="utf-8") as f:
            json.dump(person, f, ensure_ascii=False, indent=4)


def cetakExcel(cid, data, formula=False):
    from openpyxl import Workbook
    wb = Workbook()  # (write_only=True)
    ws = wb.create_sheet()
    sheet = wb.active
    for i in data:
        sheet.append(i)
    if formula:
        total = sheet.max_row
        jum = total + 5
        sheet['A{}'.format(jum)] = "TOTAL SANTRI TUNGGAKAN"
        sheet.merge_cells("A{0}:B{0}".format(jum))
        sheet['C{}'.format(jum)] = "=COUNTA(C2:C{})".format(total)
        sheet['D{}'.format(jum)] = "TOTAL TUNGGAKAN"
        sheet['E{}'.format(jum)] = "=SUM(E2:E{})".format(total)
    wb.save("{}.xlsx".format(cid))
